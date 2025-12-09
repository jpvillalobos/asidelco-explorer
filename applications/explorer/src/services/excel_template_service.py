"""
Excel Template Service - Load JSON data into Excel templates
"""
from typing import Dict, Any, List, Optional
import logging
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Fixed mapping from template headers to JSON field names
TEMPLATE_TO_JSON_MAP = {
    "RecordID": "record_id",
    "Proyecto": "csv_proyecto",
    "Exonerado": "csv_exonerado",
    "Area": "csv_area",
    "Obra": "csv_obra",
    "SubObra": "csv_subobra",
    "FechaProyecto": "csv_fechaproyecto",
    "Provincia": "csv_provincia",
    "Canton": "csv_canton",
    "Distrito": "csv_distrito",
    "Unidad": "csv_unidad",
    "Clasificacion": "csv_clasificacion",
    
    # Project-level fields
    "Estado Proyecto": "project_estado",
    "Detalle Estado": "project_detalle_de_estado",
    "Cedula Propietario": "project_cedula_propietario",
    "Nombre Propietario": "project_nombre_propietario",
    "Numero Apc": "project_num_apc",
    "Numero Boleta": "project_num_boleta",
    "Estado Boleta": "project_estado_de_la_boleta",
    "Codigo Catastro": "project_catastro",
    "Descripcion Proyecto": "project_descripcion_del_proyecto",
    "Direccion Exacta": "project_direccion_exacta",
    "Carnet Profesional": "project_carnet_profesional",
    "Carnet Empresa": "project_carnet_empresa",
    "Empresa Responsable": "project_responsable",
    
    # Project detail (primary source field)
    "Detalle Proyecto": "project_descripcion",
    
    "Valor Tasado": "project_tasado",
    "Constancia Recibido": "project_constancia_de_recibido",
    
    # Municipality-related fields
    "Fecha Ingreso Municipal": "project_fecha_de_ingreso_tramite_municipal",
    "Dias En Cola Municipal": "project_cantidad_de_dias_en_cola_de_revision_de_la_municipalidad",
    
    # Professional-related fields
    "Cedula Profesional": "professional_cedula",
    "Carnet Profesional Dup": "professional_carne",
    "Nombre Profesional": "professional_nombrecompleto",
    "Colegio Profesional": "professional_colegio",
    "Email Profesional Personal": "professional_correopermanente",
    "Email Profesional Laboral": "professional_correolaboral",
    "Condicion Profesional": "professional_condicion",
    "Telefono Profesional Movil": "professional_telcelular",
    "Telefono Profesional Oficina": "professional_teloficina",
    "Fax Profesional": "professional_fax",
    "Lugar Profesional": "professional_lugar",
    "Direccion Profesional": "professional_direccion"
}


class ExcelTemplateService:
    """Service for loading JSON data into Excel templates"""
    
    def __init__(self):
        self.template_map = TEMPLATE_TO_JSON_MAP
    
    def load_to_template(
        self,
        input_file: str,
        template_file: str,
        output_file: str,
        sheet_name: Optional[str] = None,
        context: Optional[object] = None
    ) -> Dict[str, Any]:
        """
        Load JSON data into an Excel template.
        
        Args:
            input_file: Path to input JSON file (array of records)
            template_file: Path to Excel template (.xltx or .xlsx)
            output_file: Path to output Excel file
            sheet_name: Optional sheet name (uses first sheet if not provided)
            context: Optional context for progress reporting
            
        Returns:
            Dict with status, count, and stats
        """
        logger.info("Starting Excel template loading")
        logger.info(f"  Input JSON: {input_file}")
        logger.info(f"  Template: {template_file}")
        logger.info(f"  Output: {output_file}")
        
        # Validate input file
        input_path = Path(input_file)
        if not input_path.exists():
            raise FileNotFoundError(f"Input JSON file not found: {input_file}")
        
        # Validate template file
        template_path = Path(template_file)
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_file}")
        
        # Load JSON data
        logger.info(f"Loading JSON data from {input_file}")
        try:
            with open(input_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
        except Exception as e:
            logger.error(f"Failed to load JSON file: {e}", exc_info=True)
            raise ValueError(f"Invalid JSON file: {e}")
        
        # Validate JSON is a list
        if not isinstance(json_data, list):
            raise ValueError(f"JSON data must be an array of objects, got {type(json_data).__name__}")
        
        total_records = len(json_data)
        logger.info(f"Loaded {total_records} records from JSON")
        
        if total_records == 0:
            logger.warning("No records found in JSON file")
            return {
                'status': 'warning',
                'message': 'No records to process',
                'output_file': output_file,
                'count': 0,
                'stats': {}
            }
        
        if context:
            context.report_progress(0, total_records, "Loading Excel template")
        
        # Load Excel template
        logger.info(f"Loading Excel template from {template_file}")
        try:
            workbook = load_workbook(template_file)
        except Exception as e:
            logger.error(f"Failed to load Excel template: {e}", exc_info=True)
            raise ValueError(f"Invalid Excel template file: {e}")
        
        # Get target sheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in template. Available: {workbook.sheetnames}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
            sheet_name = sheet.title
        
        logger.info(f"Using sheet: {sheet_name}")
        
        # Read template headers from first row
        headers = self._read_headers(sheet)
        logger.info(f"Found {len(headers)} headers in template: {list(headers.keys())[:5]}...")
        
        # Validate headers match expected template
        missing_headers = []
        for template_header in self.template_map.keys():
            if template_header not in headers:
                missing_headers.append(template_header)
        
        if missing_headers:
            logger.warning(f"Template is missing {len(missing_headers)} expected headers: {missing_headers[:5]}...")
        
        # Find first empty row (row after headers)
        start_row = 2  # Row 1 is headers, data starts at row 2
        
        # Statistics
        stats = {
            'total_records': total_records,
            'rows_written': 0,
            'fields_missing': 0,
            'errors': 0
        }
        
        # Write data rows
        logger.info(f"Writing {total_records} records to Excel")
        
        for i, record in enumerate(json_data):
            try:
                if not isinstance(record, dict):
                    logger.warning(f"Record {i} is not a dict, skipping: {type(record).__name__}")
                    stats['errors'] += 1
                    continue
                
                # Build row data based on template headers
                row_data = self._build_row_from_record(record, headers)
                
                # Count missing fields
                missing_in_record = sum(1 for val in row_data if val is None)
                stats['fields_missing'] += missing_in_record
                
                # Write row to sheet
                row_num = start_row + i
                for col_idx, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_num, column=col_idx)
                    cell.value = value
                
                stats['rows_written'] += 1
                
                # Report progress
                if context and (i + 1) % 100 == 0:  # Report every 100 records
                    context.report_progress(
                        i + 1,
                        total_records,
                        f"Written {i + 1}/{total_records} records",
                        {"rows_written": stats['rows_written']}
                    )
                
            except Exception as e:
                logger.error(f"Error processing record {i}: {e}", exc_info=True)
                stats['errors'] += 1
        
        # Ensure output directory exists
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Force .xlsx extension
        if output_path.suffix.lower() not in ['.xlsx']:
            output_file = str(output_path.with_suffix('.xlsx'))
            output_path = Path(output_file)
            logger.warning(f"Changed output extension to .xlsx: {output_file}")
        
        # Save workbook
        logger.info(f"Saving Excel workbook to {output_file}")
        try:
            # Remove existing file if present
            if output_path.exists():
                output_path.unlink()
                logger.debug(f"Removed existing output file")
            
            # CRITICAL FIX: Set template flag to False to save as regular workbook
            workbook.template = False
            
            # Save as regular .xlsx workbook (not template)
            workbook.save(output_file)
            
            # Close workbook to free resources
            workbook.close()
            
            logger.info(f"Successfully saved Excel file: {output_file}")
            
            # Verify file creation
            if not output_path.exists():
                raise IOError(f"Output file was not created: {output_file}")
            
            file_size = output_path.stat().st_size
            logger.info(f"Output file size: {file_size / 1024:.2f} KB")
            
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}", exc_info=True)
            raise IOError(f"Failed to save Excel file: {e}")
        
        logger.info("Excel template loading completed")
        logger.info(f"  Rows written: {stats['rows_written']}")
        logger.info(f"  Fields missing: {stats['fields_missing']}")
        logger.info(f"  Errors: {stats['errors']}")
        
        if context:
            context.report_progress(
                total_records,
                total_records,
                "Excel template loading complete",
                stats
            )
        
        return {
            'status': 'success',
            'output_file': output_file,
            'count': stats['rows_written'],
            'stats': stats
        }
    
    def _read_headers(self, sheet: Worksheet) -> Dict[str, int]:
        """
        Read headers from first row of sheet.
        
        Args:
            sheet: Excel worksheet
            
        Returns:
            Dict mapping header names to column indices (1-based)
        """
        headers = {}
        
        # Read first row
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                header_name = str(cell.value).strip()
                headers[header_name] = col_idx
        
        return headers
    
    def _build_row_from_record(
        self,
        record: Dict[str, Any],
        headers: Dict[str, int]
    ) -> List[Any]:
        """
        Build Excel row data from JSON record using template headers.
        
        Args:
            record: JSON record (dict)
            headers: Dict mapping header names to column indices
            
        Returns:
            List of values in correct column order
        """
        # Create row array with None values
        max_col = max(headers.values()) if headers else 0
        row_data = [None] * max_col
        
        # Fill in values based on template mapping
        for header_name, col_idx in headers.items():
            # Look up JSON field name
            json_field = self.template_map.get(header_name)
            
            if json_field:
                # Get value from record
                value = record.get(json_field)
                
                # Convert None to empty string for Excel
                if value is None:
                    value = ""
                
                # Handle special cases
                if isinstance(value, (list, dict)):
                    # Convert complex types to JSON string
                    value = json.dumps(value, ensure_ascii=False)
                
                # Store value at correct column index (0-based array)
                row_data[col_idx - 1] = value
            else:
                # Header not in mapping - leave blank
                logger.debug(f"Header '{header_name}' not in mapping, leaving blank")
                row_data[col_idx - 1] = ""
        
        return row_data
    
    def validate_template(
        self,
        template_file: str,
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Validate that a template file has expected headers.
        
        Args:
            template_file: Path to Excel template
            sheet_name: Optional sheet name
            
        Returns:
            Dict with validation results
        """
        logger.info(f"Validating template: {template_file}")
        
        # Load template
        try:
            workbook = load_workbook(template_file)
        except Exception as e:
            return {
                'valid': False,
                'error': f"Failed to load template: {e}"
            }
        
        # Get sheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                return {
                    'valid': False,
                    'error': f"Sheet '{sheet_name}' not found",
                    'available_sheets': workbook.sheetnames
                }
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        # Read headers
        headers = self._read_headers(sheet)
        
        # Check for expected headers
        expected_headers = set(self.template_map.keys())
        found_headers = set(headers.keys())
        
        missing_headers = expected_headers - found_headers
        extra_headers = found_headers - expected_headers
        
        return {
            'valid': len(missing_headers) == 0,
            'total_headers': len(headers),
            'expected_headers': len(expected_headers),
            'missing_headers': list(missing_headers),
            'extra_headers': list(extra_headers),
            'sheet_name': sheet.title
        }